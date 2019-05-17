import openpyxl

path = "C:/Users/khan/Desktop/QA/Data driven testing/DDT.xlsx"

workbook = openpyxl.load_workbook(path)

# First Method
sheet = workbook.active

#Second Method
#sheet = workbook.get_sheet_by_name("Sheet1")

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row = r, column = c).value, end="       ")
    print()